r"""
iERP_S9 재고현황조회(PM80160Rv1)·품목정보조회(PM10105Rv1) 엑셀 추출 자동화
⚠️ Windows 전용 + **관리자 권한 PowerShell**에서 실행 (iERP가 관리자 권한이라 맞춰야 함)

★ 전제:
  - 사용자가 **이미 iERP에 로그인**하고, 추출할 **화면(재고현황조회 등)을 띄워둔** 상태.
    (자동 로그인/자동 실행 없음 — 띄워진 창에 연결만 한다.)
  - iERP = .NET WinForms → pywinauto **backend="uia"** 사용 (정찰로 확인).

흐름(화면별): 창 연결 → [조회 조건 설정] → '조회' 클릭 → '엑셀출력' 클릭 → 저장.
그 export 파일을 run.py 파이프라인에 넘긴다.

의존성:  py -m pip install pywinauto openpyxl
"""
from __future__ import annotations
import datetime
import time
from pathlib import Path


def _ts() -> str:
    """저장일시 고유값 (YYYYMMDD_HHMMSS)."""
    return datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

try:
    from pywinauto.application import Application
    from pywinauto.keyboard import send_keys
except ImportError:
    Application = None

BACKEND = "uia"
EXPORT_DIR = Path(r"C:\ierp_exports")

# 정찰(inspect_ierp.py)로 확인된 화면별 식별자
STOCK = dict(
    title_re=r".*재고현황조회.*",       # frmPM80160Rv1
    program_id="PM80160Rv1",           # F9 실행창에 입력할 프로그램 ID(풀네임)
    item_type_combo="cmbMTYPN",        # 품목유형(필수 선택)
    item_type_value="ALL",             # ALL = 전체 유형 한번에
    item_type_index=0,                 # ALL = 드롭다운 1번째
    query_btn="조회",
    export_btn="엑셀출력",
)
BOM = dict(
    title_re=r".*품목정보조회.*",       # frmPM10105Rv1
    program_id="PM10105Rv1",           # F9 실행창에 입력할 프로그램 ID(풀네임)
    menu_path=["자원정의", "품목정의"],  # iEMenu 트리 경로(F9 실패 시 폴백)
    menu_program="품목정보조회",         # 우측 목록에서 더블클릭할 프로그램
    mode_combo="cmbSELECT",            # 조회 모드 드롭다운
    mode_value="다단계전개",            # 다단계 BOM 전개
    mode_index=2,                      # 품목코드(0)/1단계전개(1)/다단계전개(2)
    item_combo="cmbMITEM",             # 품목코드 입력 콤보
    auto_query=True,                   # 품목코드 입력 시 자동조회 → '조회' 버튼 클릭 생략
    query_wait=5.0,                    # 자동조회 로딩 대기(초)
    query_btn="조회",
    export_btn="엑셀출력",
)


def connect_window(title_re: str):
    """이미 열려 있는 iERP 화면 창에 연결. 없으면 안내 후 중단."""
    if Application is None:
        raise SystemExit("pywinauto 미설치 — 'py -m pip install pywinauto' 후 실행하세요.")
    try:
        app = Application(backend=BACKEND).connect(title_re=title_re, timeout=8)
    except Exception:
        raise SystemExit(
            f"창({title_re})을 찾지 못했습니다.\n"
            "→ iERP에 로그인했는지, 메뉴(iEMenu)가 떠 있는지 확인 후 다시 실행하세요.")
    return app.window(title_re=title_re)


def _try_connect(title_re: str):
    """창이 있으면 반환, 없으면 None."""
    try:
        app = Application(backend=BACKEND).connect(title_re=title_re, timeout=3)
        return app.window(title_re=title_re)
    except Exception:
        return None


def open_via_menu(menu_path: list[str], program_name: str):
    """iEMenu 트리에서 menu_path 노드들을 클릭 → 우측 목록의 program_name 더블클릭(화면 열기)."""
    app = Application(backend=BACKEND).connect(title_re=r".*iEMenu.*", timeout=8)
    menu = app.window(title_re=r".*iEMenu.*")
    menu.set_focus()
    # 툴바 '전체 펼치기'로 트리 모두 펼침(더블클릭 토글 문제 회피)
    try:
        menu.child_window(title="전체 펼치기", control_type="Button").click_input()
        time.sleep(0.6)
    except Exception:
        pass
    tree = menu.child_window(auto_id="mainTreeView", control_type="Tree")
    # ⚠️ '전체 펼치기' 후 트리가 아래로 스크롤되면 상위 메뉴가 화면 밖으로 가려져
    #    click_input(화면 좌표 클릭)이 빗나가 메뉴가 안 열린다 → 트리를 맨 위로 스크롤.
    try:
        tree.set_focus()
        send_keys("^{HOME}")        # 트리 맨 위로 (Ctrl+Home)
        time.sleep(0.3)
    except Exception:
        pass
    # 마지막 카테고리(예: 품목정의) 단일 클릭 → 우측에 프로그램 목록 표시
    target = tree.child_window(title=menu_path[-1], control_type="TreeItem")
    try:
        target.scroll_into_view()   # 지원되면 대상 항목을 화면 안으로 (미지원 시 무시)
    except Exception:
        pass
    target.click_input()
    time.sleep(0.6)
    menu.child_window(title=program_name).double_click_input()  # 우측 목록 프로그램
    time.sleep(2.0)


def open_via_f9(program_id: str):
    """iEMenu 메인창 포커스 상태에서 F9 → 화면 중앙에 뜨는 실행창에
    프로그램 ID(풀네임)를 입력하고 Enter → 해당 화면을 바로 연다.
    (기존 트리 '전체 펼치기' 탐색보다 빠르고 정확)."""
    app = Application(backend=BACKEND).connect(title_re=r".*iEMenu.*", timeout=8)
    menu = app.window(title_re=r".*iEMenu.*")
    menu.set_focus()
    time.sleep(0.3)
    send_keys("{F9}")               # 실행창 팝업 (화면 중앙 작은 창)
    time.sleep(0.8)                 # 팝업이 뜨고 입력 포커스가 잡힐 때까지 대기
    # 팝업이 뜨면 입력칸에 포커스가 있으므로 그대로 타이핑 → Enter.
    # 프로그램 ID에 특수문자 없다고 보고 그대로 전송(오타 방지 위해 기존값 지우고 입력).
    send_keys("^a{DELETE}")         # 혹시 남은 값 전체선택 후 삭제
    send_keys(program_id, with_spaces=True)
    time.sleep(0.2)
    send_keys("{ENTER}")
    time.sleep(2.0)                 # 화면 로딩 대기


def ensure_screen_open(cfg: dict):
    """화면이 열려 있으면 그 창, 아니면 F9 실행창으로 자동으로 열고 연결.
    (F9 실패 시 기존 트리 메뉴 탐색으로 폴백)."""
    win = _try_connect(cfg["title_re"])
    if win is not None:
        return win
    # 1순위: F9 실행창에 프로그램 ID 입력
    if cfg.get("program_id"):
        print(f"   화면 자동 열기(F9): {cfg['program_id']}")
        try:
            open_via_f9(cfg["program_id"])
            win = _try_connect(cfg["title_re"])
            if win is not None:
                return win
            print("   F9 열기 실패 → 트리 메뉴 탐색으로 폴백")
        except Exception as e:
            print(f"   F9 열기 예외({type(e).__name__}) → 트리 메뉴 탐색으로 폴백")
    # 2순위(폴백): iEMenu 트리 탐색
    if cfg.get("menu_path"):
        print(f"   화면 자동 열기: {' > '.join(cfg['menu_path'])} > {cfg['menu_program']}")
        open_via_menu(cfg["menu_path"], cfg["menu_program"])
        return connect_window(cfg["title_re"])
    return connect_window(cfg["title_re"])  # 경로 없으면 기존 동작(안내 후 중단)


def _click(win, title: str):
    win.child_window(title=title, control_type="Button").click_input()


def _select_combo(win, auto_id: str, value: str, index: int | None = None) -> bool:
    """WinForms 콤보 값 선택. 인덱스가 있으면 키보드로 결정적 선택(권장),
    없으면 펼쳐서 목록 항목(정확 일치)을 클릭."""
    combo = win.child_window(auto_id=auto_id, control_type="ComboBox")
    if index is not None:                  # 결정적: 드롭다운 열고 처음으로 → index만큼 아래 → 확정
        try:
            combo.click_input()
            time.sleep(0.4)
            send_keys("{HOME}")
            for _ in range(index):
                send_keys("{DOWN}")
            send_keys("{ENTER}")
            time.sleep(0.3)
            return True
        except Exception:
            pass
    try:                                   # 폴백: 펼치고 목록 항목 클릭
        combo.click_input()
        time.sleep(0.4)
        for scope in (combo, win):
            try:
                scope.child_window(title=value, control_type="ListItem").click_input()
                return True
            except Exception:
                continue
    except Exception:
        pass
    return False


def _excel_apps(win32, pythoncom):
    """실행 중인 모든 Excel.Application 인스턴스(ROT 열거). 실패 시 GetActiveObject 폴백."""
    apps, seen = [], set()
    try:
        rot = pythoncom.GetRunningObjectTable()
        ctx = pythoncom.CreateBindCtx(0)
        for mon in rot:
            try:
                nm = mon.GetDisplayName(ctx, None) or ""
            except Exception:
                continue
            low = nm.lower()
            if "excel" not in low and not low.endswith((".xlsx", ".xls", ".xlsm")):
                continue
            try:
                disp = win32.Dispatch(rot.GetObject(mon))
                app = getattr(disp, "Application", disp)
                h = int(app.Hwnd)
                if h not in seen:
                    seen.add(h)
                    apps.append(app)
            except Exception:
                continue
    except Exception:
        pass
    if not apps:
        try:
            apps.append(win32.GetActiveObject("Excel.Application"))
        except Exception:
            pass
    return apps


def _wait_query_done(win, grid_auto_id="dataGridView1", timeout=30,
                     stable_needed=3, poll=0.4):
    """조회 후 그리드 행수가 '늘다가 멈출 때(=로딩 완료)'까지 대기.

    ⚠️ item_count>0(첫 행 등장)만 보고 엑셀출력을 누르면, iERP가 아직 데이터를
    채우는 중에 funcCrtCSV가 돌아 'HRESULT 0x800A01A8(개체 필요)'이 난다.
    → 행수가 stable_needed 번 연속 그대로일 때만 통과한다(고정 sleep보다 빠르고 안전).
    """
    try:
        grid = win.child_window(auto_id=grid_auto_id, control_type="Table")
    except Exception:
        time.sleep(1.5)
        return
    end = time.time() + timeout
    prev, stable = None, 0
    while time.time() < end:
        try:
            cnt = grid.item_count()
        except Exception:
            time.sleep(1.5)        # 카운트 미지원 → 폴백 후 종료
            return
        if cnt == prev:            # 직전 폴링과 행수 동일 → 더 안 늘어남
            stable += 1
            if stable >= stable_needed:
                return
        else:                      # 아직 행이 늘고 있음 → 로딩 중
            prev, stable = cnt, 0
        time.sleep(poll)


def _find_export_wb(xl):
    """저장 안 됐고(미저장) 데이터가 있는 워크북 1개 반환(없으면 None)."""
    for i in range(xl.Workbooks.Count, 0, -1):
        wb = xl.Workbooks(i)
        try:
            if wb.Path:                   # 이미 저장된 것은 건너뜀
                continue
            used = wb.Worksheets(1).UsedRange
            if used.Rows.Count < 2 and used.Columns.Count < 2:
                continue                  # 아직 비어있음
            return wb
        except Exception:
            continue
    return None


def _save_active_excel(out_path: Path, timeout=120,
                       settle=6.0, stable_needed=3, poll=1.0) -> bool:
    """'엑셀출력'이 데이터를 새 엑셀로 열면, 방금 출력된 워크북을 찾아 저장하고 닫는다.

    ⚠️ iERP의 funcCrtCSV가 셀을 **다 채우기 전에** 워크북을 Close/Quit 하면 iERP가
    쓰던 COM 객체가 사라져 'HRESULT 0x800A01A8(개체 필요)' 에러가 난다.
    → ① settle 초만큼 기다린 뒤 ② UsedRange 크기가 stable_needed 번 연속 그대로일 때
       (= funcCrtCSV 쓰기 완료)만 저장한다. 쓰는 중에는 절대 건드리지 않는다.
    """
    try:
        import win32com.client as win32  # pywin32
        import pythoncom
    except ImportError:
        print("※ pywin32 미설치 — 'py -m pip install pywin32' 후 자동저장됩니다.")
        return False
    out_path.parent.mkdir(parents=True, exist_ok=True)
    time.sleep(settle)                    # iERP가 Excel 띄우고 쓰기 시작할 시간
    end = time.time() + timeout
    prev_sig, stable = None, 0
    while time.time() < end:
        for xl in _excel_apps(win32, pythoncom):
            try:
                wb = _find_export_wb(xl)
                if wb is None:
                    continue
                used = wb.Worksheets(1).UsedRange
                sig = (wb.Name, used.Rows.Count, used.Columns.Count)
                if sig == prev_sig:       # 직전 폴링과 크기 동일 → 쓰기 멈춤
                    stable += 1
                else:                     # 아직 행/열이 늘고 있음 → funcCrtCSV 쓰는 중
                    prev_sig, stable = sig, 0
                if stable < stable_needed:
                    break                 # 안정화 전엔 Close 금지, 다음 폴링까지 대기
                xl.DisplayAlerts = False
                wb.SaveAs(str(out_path), FileFormat=51)   # 51 = .xlsx
                wb.Close(SaveChanges=False)
                try:
                    if xl.Workbooks.Count == 0:
                        xl.Quit()
                except Exception:
                    pass
                return True
            except Exception:             # COM busy 등 → 백오프 후 재시도
                pass
        time.sleep(poll)
    return False


def _strip_raw_materials(path: Path) -> None:
    """저장된 BOM export(xlsx)에서 '원료' 행을 모두 삭제.
    판별: 품목유형 컬럼값='원료' 또는 품목코드가 6+6자리(6xxxxxx). 실패해도 조용히 통과."""
    try:
        import re
        import openpyxl
    except ImportError:
        return
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb.worksheets[0]
        hdr = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
        type_idxs = [i for i, h in enumerate(hdr) if h == "품목유형"]
        ci_type = type_idxs[1] if len(type_idxs) >= 2 else (type_idxs[0] if type_idxs else None)
        ci_code = next((i for i, h in enumerate(hdr) if h == "품목코드"), None)
        removed = 0
        for r in range(ws.max_row, 1, -1):     # 아래→위로 삭제(행 밀림 방지)
            typ = ws.cell(row=r, column=ci_type + 1).value if ci_type is not None else None
            code = ws.cell(row=r, column=ci_code + 1).value if ci_code is not None else None
            is_raw = (str(typ).strip() == "원료"
                      or bool(re.match(r"^6\d{6}$", str(code or "").strip())))
            if is_raw:
                ws.delete_rows(r, 1)
                removed += 1
        if removed:
            wb.save(path)
    except Exception:
        pass


def export_screen(cfg: dict, out_path: Path, item_code: str | None = None) -> Path:
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    win = ensure_screen_open(cfg)       # 없으면 메뉴로 자동 열기
    win.set_focus()
    time.sleep(0.5)
    if cfg.get("mode_combo"):           # 조회 모드(예: 다단계전개)
        if not _select_combo(win, cfg["mode_combo"], cfg["mode_value"], cfg.get("mode_index")):
            print(f"※ 조회모드 선택 실패. 화면에서 '{cfg['mode_value']}' 직접 선택하세요.")
        time.sleep(0.3)
    if cfg.get("item_combo") and item_code:   # 품목코드 입력
        try:
            combo = win.child_window(auto_id=cfg["item_combo"], control_type="ComboBox")
            try:
                combo.child_window(control_type="Edit").set_text(item_code)
            except Exception:
                combo.set_edit_text(item_code)
            send_keys("{ENTER}")
        except Exception as e:
            print(f"※ 품목코드 입력 실패({e}).")
        time.sleep(0.3)
    if cfg.get("item_type_combo"):     # 품목유형 선택(재고화면)
        if not _select_combo(win, cfg["item_type_combo"], cfg["item_type_value"], cfg.get("item_type_index")):
            print(f"※ 품목유형 자동선택 실패. 직접 '{cfg['item_type_value']}' 선택하세요.")
        time.sleep(0.5)
    if cfg.get("auto_query"):          # 품목코드 입력 시 자동조회되는 화면 → 조회 버튼 생략
        time.sleep(cfg.get("query_wait", 1.5))   # 자동조회 로딩 대기
    else:
        _click(win, cfg["query_btn"])  # 조회
        _wait_query_done(win, cfg.get("grid", "dataGridView1"))   # 조회 완료(그리드 채워짐) 확인
    _click(win, cfg["export_btn"])     # 엑셀출력 → 엑셀로 데이터 열림
    if _save_active_excel(out_path):   # 엑셀에 데이터 다 뜬 것 확인 후 저장
        _strip_raw_materials(out_path)  # 저장 파일에서 '원료' 행 제거
        print(f"저장됨: {out_path}")
    else:
        print(f"※ 자동저장 실패 — 열린 엑셀을 '{out_path}' 로 직접 저장하세요.")
    return out_path


def export_stock(out_path: Path | None = None) -> Path:
    return export_screen(STOCK, out_path or (EXPORT_DIR / f"PM80160Rv1_stock_{_ts()}.xlsx"))


def export_bom(item_code: str, out_path: Path | None = None) -> Path:
    """단일 완제품의 다단계전개 BOM export. 여러 품목은 반복 호출."""
    out = out_path or (EXPORT_DIR / f"PM10105Rv1_bom_{item_code}_{_ts()}.xlsx")
    return export_screen(BOM, out, item_code=item_code)


def export_boms(item_codes) -> list[Path]:
    """여러 완제품의 BOM을 순차 추출(각 파일명에 저장일시). 반환: 저장 경로 목록."""
    paths = []
    for code in item_codes:
        try:
            paths.append(export_bom(code))
            time.sleep(1)
        except Exception as e:
            print(f"※ {code} BOM 추출 실패: {e}")
    return paths


if __name__ == "__main__":
    # ★ 관리자 권한 PowerShell에서, iERP 로그인 + 해당 화면 띄운 상태로 실행.
    #   재고현황조회 화면을 띄워두고:  py ierp_export.py
    p = export_stock()
    print("재고 export 시도:", p)
    print(f"다음:  py run.py --bom <BOM.xlsx> --stock {p}")
