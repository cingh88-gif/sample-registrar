r"""
초도분 추출 결과 → 표준견본 웹앱용 register.json 생성.

웹 스키마(필수 customer/code/name; 나머지는 웹이 채움):
  [{"customer","code","name","category","supply","regDate"}, ...]

사용:
  # 화면 '엑셀출력' 결과에 초도입고일/자재입고일 컬럼이 있는 경우 (오프라인 가능, 권장)
  py sample_register.py --excel C:\ierp_exports\receipt_xxxx.xlsx --out register.json
  py sample_register.py --excel receipt.xlsx --subsidiary-only      # 부자재(접미사 7/5)만

  # 화면 자동화(우클릭) 경로 — RECEIPT 정찰 완료 후
  py sample_register.py --screen --from 2026-06-19 [--to 2026-06-19]

생성된 register.json 을 표준견본 웹앱의 'iERP 등록 가져오기' 버튼으로 업로드한다.
※ 실데이터 register.json 은 공개 저장소/배포에 올리지 말 것(로컬에서만 사용).
"""
from __future__ import annotations
import argparse
import datetime
import json
from pathlib import Path

import customer_map as cm
import sample_export as se


def _today() -> str:
    return datetime.date.today().strftime("%Y-%m-%d")


def map_row_to_record(row: dict) -> dict | None:
    """초도분 행 1건 → 웹 레코드. 필수값(code/name) 없으면 None."""
    code = str(row.get("code") or "").strip()
    name = str(row.get("name") or "").strip()
    if not code or not name:
        return None

    customer = str(row.get("customer") or "").strip()
    if not customer:                                   # 화면에 고객사 없으면 코드로 도출
        customer, matched = cm.customer_from_code(code)
        if not matched:
            print(f"  ※ 거래처 코드 미등록: {code[:4]} (코드 {code}) → 고객사='{customer}' 그대로 사용")

    category = str(row.get("category") or "").strip() or cm.form_from_code(code)
    # 사급/자급: 명시값이 사급/자급이면 그대로, 아니면 입고유형 텍스트로 도출(사급반입→사급, 그 외→자급)
    sup_src = str(row.get("supply") or "").strip()
    if sup_src in ("사급", "자급"):
        supply = sup_src
    elif "사급" in sup_src:           # 예: '사급반입'
        supply = "사급"
    else:                             # 구매입고/외주입고/빈값 등
        supply = "자급"
    reg_date = se._norm_date(row.get("mat_in")) or _today()

    return {
        "customer": customer,
        "code": code,
        "name": name,
        "category": category,
        "supply": supply,
        "regDate": reg_date,
    }


def build_register(rows: list[dict], subsidiary_only: bool = False,
                   chodo_from: str | None = None, chodo_to: str | None = None,
                   min_as_first: bool = False) -> list[dict]:
    """행 목록 → 초도분 필터 → (옵션) 초도입고일 범위/부자재 → 웹 레코드 목록.

    초도입고일은 제품마다 iERP에 저장된 고유값이다(우클릭 등으로 취득해 first_in 에 채워야 함).
    --min-as-first(min_as_first=True) 를 주면 '입고이력 최초일'로 추정하지만, 이는 정확하지 않다."""
    has_first = any(se._norm_date(r.get("first_in")) for r in rows)
    if not has_first:
        if not min_as_first:
            raise SystemExit(
                "초도입고일 값이 없습니다.\n"
                "  초도입고일은 제품별 iERP 저장값(우클릭으로 보이는 값)이라, 입고일자만으로는 알 수 없습니다.\n"
                "  → 초도입고일을 가져오는 경로(우클릭 자동화 또는 초도입고일이 있는 화면/리포트)가 필요합니다.\n"
                "  (정확도를 포기하고 '입고이력 최초일'로 추정만 해보려면 --min-as-first)")
        lo, hi = se.history_date_range(rows)
        print(f"※ [추정] 초도입고일 컬럼 없음 → 입고이력 최초일로 추정. 엑셀 입고일 범위: {lo} ~ {hi}")
        if lo == hi and lo:
            print("  ⚠️ 입고일이 하루뿐 — 전체 이력이 아니면 과대 집계됩니다.")
        se.fill_first_in_from_history(rows)

    keep = se.chodo_rows(rows)
    if chodo_from or chodo_to:              # 초도입고일(=등록 대상일) 범위 필터
        def inrange(r):
            d = se._norm_date(r.get("first_in"))
            return (not chodo_from or d >= chodo_from) and (not chodo_to or d <= chodo_to)
        keep = [r for r in keep if inrange(r)]
    if subsidiary_only:
        keep = [r for r in keep if cm.is_subsidiary(r.get("code"))]
    out, skipped = [], 0
    seen = set()
    for r in keep:
        rec = map_row_to_record(r)
        if rec is None:
            skipped += 1
            continue
        if rec["code"] in seen:                        # 파일 내 중복 코드 제거(웹도 다시 거름)
            continue
        seen.add(rec["code"])
        out.append(rec)
    print(f"초도분 {len(keep)}건 → 등록대상 {len(out)}건"
          + (f" (필수값 누락 {skipped}건 제외)" if skipped else ""))
    return out


def analyze(rows: list[dict], dedup_same_day: bool = True) -> list[dict]:
    """행 목록 → 화면/엑셀 표시용 분석 결과. (first_in 은 이미 채워져 있어야 함)
    각 항목: customer, code, name, category, supply, regDate, mat_in(입고일), first_in(초도입고), is_chodo.

    dedup_same_day: 같은 입고일에 같은 품목코드가 여러 번(여러 전표/로트) 들어온 중복 행을 1건으로 합침."""
    out, seen = [], set()
    for r in rows:
        rec = map_row_to_record(r)            # customer/category/supply/regDate 도출 (code/name 필수)
        if rec is None:
            continue
        mat = se._norm_date(r.get("mat_in"))
        first = se._norm_date(r.get("first_in"))
        if dedup_same_day:
            key = (rec["code"], mat)          # 품목코드 + 입고일 동일 = 중복
            if key in seen:
                continue
            seen.add(key)
        out.append({**rec, "mat_in": mat, "first_in": first,
                    "is_chodo": bool(first) and first == mat})
    return out


def records_from_analysis(analysis: list[dict], subsidiary_only: bool = False) -> list[dict]:
    """분석 결과 → 초도분만, 웹 등록 레코드(customer/code/name/category/supply/regDate)."""
    out, seen = [], set()
    for a in analysis:
        if not a.get("is_chodo"):
            continue
        if subsidiary_only and not cm.is_subsidiary(a["code"]):
            continue
        if a["code"] in seen:
            continue
        seen.add(a["code"])
        out.append({k: a[k] for k in ("customer", "code", "name", "category", "supply", "regDate")})
    return out


def write_review_excel(analysis: list[dict], out_path: Path) -> Path:
    """분석 결과(전체 + 초도분 여부)를 검토용 엑셀로 저장."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "초도분분석"
    ws.append(["초도분", "품목코드", "품목명", "고객사", "유형", "사/자급", "입고일", "초도입고일"])
    red = Font(color="FF0000", bold=True)
    for a in sorted(analysis, key=lambda x: (not x["is_chodo"], x["code"])):
        first = a["first_in"]
        manual = not first                       # 초도입고 못 읽음 → 수동조회요망
        ws.append(["O" if a["is_chodo"] else ("수동" if manual else "X"),
                   a["code"], a["name"], a["customer"], a["category"], a["supply"],
                   a["mat_in"], "수동조회요망" if manual else first])
        if manual:                               # 초도분 / 초도입고일 칸 빨간 글씨
            r = ws.max_row
            ws.cell(row=r, column=1).font = red
            ws.cell(row=r, column=8).font = red

    # 헤더 행 서식: 음영 #002060, 굵게, 흰색, 가운데정렬
    header_fill = PatternFill("solid", fgColor="002060")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    # 열너비
    widths = {"A": 6, "B": 13, "C": 80, "D": 28, "E": 20, "F": 6, "G": 11, "H": 11}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    n_chodo = sum(1 for a in analysis if a["is_chodo"])
    n_manual = sum(1 for a in analysis if not a["first_in"])
    print(f"검토 엑셀 저장: {out_path}  (전체 {len(analysis)} · 초도분 {n_chodo}"
          + (f" · 수동조회요망 {n_manual}" if n_manual else "") + ")")
    return out_path


def write_register_json(records: list[dict], out_path: Path) -> Path:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"저장됨: {out_path}  ({len(records)}건)")
    return out_path


def _default_out() -> Path:
    base = se.ie.EXPORT_DIR if (se.ie is not None) else Path(".")
    try:
        ts = se.ie._ts() if se.ie is not None else datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    except Exception:
        ts = "out"
    return base / f"run_{ts}" / "register.json"


def main():
    ap = argparse.ArgumentParser(description="iERP 초도분 → 표준견본 register.json")
    src = ap.add_mutually_exclusive_group(required=True)
    src.add_argument("--excel", type=Path, help="부자재 입고 엑셀출력(.xlsx) — 초도입고일/자재입고일 컬럼 포함")
    src.add_argument("--screen", action="store_true", help="화면 자동화(우클릭) 경로 (정찰 완료 필요)")
    ap.add_argument("--from", dest="date_from", help="초도입고일 시작 YYYY-MM-DD (--excel: 등록대상 범위 / --screen: 조회 시작일)")
    ap.add_argument("--to", dest="date_to", help="초도입고일 종료 YYYY-MM-DD")
    ap.add_argument("--on", dest="date_on", help="특정 초도입고일 하루만 (= --from X --to X)")
    ap.add_argument("--subsidiary-only", action="store_true", help="부자재 계열(접미사 7/5)만 등록")
    ap.add_argument("--lookup-firstin", action="store_true",
                    help="품목재고조회(ITEMINVv1) 창에서 코드별 초도입고일을 읽어와 채움 (iERP 화면 띄운 상태)")
    ap.add_argument("--min-as-first", action="store_true",
                    help="[비권장] 초도입고일을 입고이력 최초일로 추정(부정확). 정식은 품목재고조회 초도입고값 사용")
    ap.add_argument("--out", type=Path, help="출력 register.json 경로")
    args = ap.parse_args()

    cfrom = args.date_on or args.date_from
    cto = args.date_on or args.date_to

    if args.excel:
        rows = se.load_receipt_excel(args.excel)
        print(f"엑셀 {args.excel} → {len(rows)}행")
        if args.lookup_firstin:        # 품목재고조회에서 코드별 초도입고일 조회 → first_in 채움
            print("품목재고조회(ITEMINVv1)에서 초도입고일 조회 중…")
            fmap = se.lookup_first_in([r.get("code") for r in rows])
            for r in rows:
                r["first_in"] = fmap.get(str(r.get("code") or "").strip(), "")
        records = build_register(rows, subsidiary_only=args.subsidiary_only,
                                 chodo_from=cfrom, chodo_to=cto, min_as_first=args.min_as_first)
    else:
        if not args.date_from:
            ap.error("--screen 사용 시 --from(입고일자 시작)이 필요합니다.")
        # 품질검사대상조회 자동 엑셀출력 → 초도입고 조회 → 초도분
        xlsx = se.export_receipt(args.date_from, args.date_to)
        rows = se.load_receipt_excel(xlsx)
        print(f"엑셀 {xlsx} → {len(rows)}행")
        fmap = se.lookup_first_in([r.get("code") for r in rows])
        for r in rows:
            r["first_in"] = fmap.get(str(r.get("code") or "").strip(), "")
        records = build_register(rows, subsidiary_only=args.subsidiary_only,
                                 chodo_from=cfrom, chodo_to=cto)
    out = args.out or _default_out()
    write_register_json(records, out)


if __name__ == "__main__":
    main()
